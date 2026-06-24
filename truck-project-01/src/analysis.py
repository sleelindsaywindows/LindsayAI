"""
Excel report generator for Lindsay Windows Load Planner.
Produces a 4-sheet workbook: Summary, Per-Route Detail, Cost Comparison, Constraint Audit.
"""

import sys
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from .models import TruckAssignment, Order

# Fill colours
_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_RED    = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
_GREEN  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_row(ws, values: list, row: int = 1):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _HEADER
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _sheet_summary(wb, assignments: list, dropped: list, josephs_truck_count: int):
    ws = wb.create_sheet("Summary")
    headers = [
        "Truck Name", "Type", "Stops",
        "Sq Ft Used", "Capacity (sq ft)", "Utilization %",
        "Est. Miles", "Est. Cost ($)",
    ]
    _header_row(ws, headers, 1)

    total_sqft = 0.0
    total_miles = 0.0
    total_cost = 0.0
    util_values = []

    for r, a in enumerate(assignments, 2):
        miles = a.route_distance_miles if a.route_distance_miles else 0.0
        cost = (miles * a.truck.cost_per_mile) if miles else 0.0
        util = round(a.utilization_pct, 1)
        util_values.append(util)
        total_sqft += a.total_capacity_used
        total_miles += miles
        total_cost += cost

        row_data = [
            a.truck.name,
            a.truck.truck_type,
            len(a.stops),
            round(a.total_capacity_used, 1),
            round(a.truck.max_capacity, 1),
            util,
            round(miles, 1) if miles else "N/A",
            round(cost, 2) if miles else "N/A",
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER

    # Summary row
    last_data_row = len(assignments) + 1
    summary_row = last_data_row + 2
    avg_util = round(sum(util_values) / len(util_values), 1) if util_values else 0.0
    ws.cell(row=summary_row, column=1, value="TOTALS / SUMMARY").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=round(total_sqft, 1)).font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=avg_util).font = Font(bold=True)
    if total_miles:
        ws.cell(row=summary_row, column=7, value=round(total_miles, 1)).font = Font(bold=True)
        ws.cell(row=summary_row, column=8, value=round(total_cost, 2)).font = Font(bold=True)
    ws.cell(
        row=summary_row + 1, column=1,
        value=f"Optimizer trucks used: {len(assignments)}  |  Joseph's trucks: {josephs_truck_count}"
    ).font = Font(italic=True)
    if dropped:
        ws.cell(row=summary_row + 2, column=1,
                value=f"Dropped orders: {len(dropped)}").font = Font(bold=True, color="FF0000")

    # Bar chart: Utilization % per truck
    if assignments:
        chart1 = BarChart()
        chart1.type = "bar"  # horizontal
        chart1.title = "Utilization % per Truck"
        chart1.y_axis.title = "Truck"
        chart1.x_axis.title = "Utilization %"
        chart1.shape = 4
        data_ref = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=len(assignments) + 1)
        cats_ref = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=len(assignments) + 1)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.width = 20
        chart1.height = max(8, len(assignments) * 0.8)
        ws.add_chart(chart1, f"A{summary_row + 4}")

        # Bar chart: Est. cost per truck (only when geocoding ran)
        if total_miles:
            chart2 = BarChart()
            chart2.type = "bar"
            chart2.title = "Estimated Cost per Truck ($)"
            data_ref2 = Reference(ws, min_col=8, max_col=8, min_row=1, max_row=len(assignments) + 1)
            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref)
            chart2.width = 20
            chart2.height = max(8, len(assignments) * 0.8)
            col_offset = get_column_letter(11)
            ws.add_chart(chart2, f"{col_offset}{summary_row + 4}")

    _auto_width(ws)


def _sheet_detail(wb, assignments: list, dropped: list):
    ws = wb.create_sheet("Per-Route Detail")
    headers = [
        "Truck", "Stop #", "Customer", "Address", "State",
        "Sq Ft", "Truck Restriction", "Priority", "Notes",
    ]
    _header_row(ws, headers, 1)

    row = 2
    for a in assignments:
        for stop in a.stops:
            o = stop.order
            state = ""
            # Grab 2-letter state token heuristic (uppercase, len 2)
            for p in reversed(o.address.split()):
                if len(p) == 2 and p.isalpha():
                    state = p.upper()
                    break

            restriction = ", ".join(o.allowed_truck_types) if o.allowed_truck_types else ""
            row_data = [
                a.truck.name,
                stop.stop_number,
                o.customer_name,
                o.address,
                state,
                round(o.capacity_units, 1),
                restriction,
                o.priority,
                o.notes,
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = _BORDER
                if o.allowed_truck_types:
                    cell.fill = _YELLOW
            row += 1

    # Dropped orders at bottom in red
    if dropped:
        ws.cell(row=row, column=1, value="DROPPED ORDERS").font = Font(bold=True, color="FF0000")
        row += 1
        for o in dropped:
            restriction = ", ".join(o.allowed_truck_types) if o.allowed_truck_types else ""
            row_data = [
                "UNASSIGNED", "", o.customer_name, o.address, "",
                round(o.capacity_units, 1), restriction, o.priority, o.notes,
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.fill = _RED
                cell.border = _BORDER
            row += 1

    _auto_width(ws)


def _sheet_cost(wb, assignments: list, dropped: list, josephs_truck_count: int):
    ws = wb.create_sheet("Cost Comparison")
    _header_row(ws, ["Metric", "Optimizer", "Current Process"], 1)

    total_miles = sum(a.route_distance_miles for a in assignments if a.route_distance_miles)
    total_cost = sum(
        a.route_distance_miles * a.truck.cost_per_mile
        for a in assignments
        if a.route_distance_miles
    )
    util_values = [a.utilization_pct for a in assignments]
    avg_util = round(sum(util_values) / len(util_values), 1) if util_values else 0.0

    has_miles = total_miles > 0
    rows = [
        ("Trucks Used", len(assignments), str(josephs_truck_count)),
        ("Total Est. Miles", round(total_miles, 1) if has_miles else "unknown", "unknown"),
        ("Total Est. Cost ($)", round(total_cost, 2) if has_miles else "unknown", "unknown"),
        ("Avg Utilization %", avg_util, "unknown"),
        ("Dropped Orders", len(dropped), "unknown"),
    ]
    for r, (metric, opt_val, joseph_val) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=metric).border = _BORDER
        ws.cell(row=r, column=2, value=opt_val).border = _BORDER
        ws.cell(row=r, column=3, value=joseph_val).border = _BORDER

    note_row = len(rows) + 3
    ws.cell(
        row=note_row, column=1,
        value="Note: Current process cost computable once the 6/17 manual route sheet is entered.",
    ).font = Font(italic=True)
    if not has_miles:
        ws.cell(
            row=note_row + 1, column=1,
            value="Mileage estimates require geocoding — enable in Load Plan tab.",
        ).font = Font(italic=True, color="FF6600")

    _auto_width(ws)


def _sheet_audit(wb, assignments: list) -> int:
    """Returns count of FAIL rows."""
    ws = wb.create_sheet("Constraint Audit")
    headers = ["Order ID", "Customer", "Assigned Truck", "Truck Type", "Result"]
    _header_row(ws, headers, 1)

    restricted_stops = [
        (a, stop)
        for a in assignments
        for stop in a.stops
        if stop.order.allowed_truck_types
    ]

    fail_count = 0
    for r, (a, stop) in enumerate(restricted_stops, 2):
        o = stop.order
        passed = a.truck.truck_type in o.allowed_truck_types
        result = "PASS" if passed else "FAIL"
        if not passed:
            fail_count += 1
        row_data = [o.order_id, o.customer_name, a.truck.name, a.truck.truck_type, result]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            cell.fill = _GREEN if passed else _RED

    summary_row = len(restricted_stops) + 3
    ws.cell(
        row=summary_row, column=1,
        value=(
            f"{len(restricted_stops)} homebuilder stop(s) checked. "
            f"{len(restricted_stops) - fail_count} passed, {fail_count} failed."
        ),
    ).font = Font(bold=True)

    _auto_width(ws)
    return fail_count


def generate_report(
    assignments: list,
    dropped: list,
    orders: list,
    output_path: str = "lindsay_analysis.xlsx",
    josephs_truck_count: int = 13,
) -> str:
    """
    Generate Excel workbook with 4 sheets:
      1. Summary — truck utilization + cost table + charts
      2. Per-Route Detail — every stop with yellow homebuilder highlight, red dropped
      3. Cost Comparison — optimizer vs current manual routing process
      4. Constraint Audit — homebuilder truck-type constraint PASS/FAIL per stop

    Returns output_path. Prints warning to stderr if any constraint FAILs.
    NOTE: josephs_truck_count defaults to 13 for the 6/17 historical run.
          Update once more than one week of data is available for a real average.
    """
    wb = openpyxl.Workbook()
    del wb["Sheet"]  # remove default empty sheet

    _sheet_summary(wb, assignments, dropped, josephs_truck_count)
    _sheet_detail(wb, assignments, dropped)
    _sheet_cost(wb, assignments, dropped, josephs_truck_count)
    fail_count = _sheet_audit(wb, assignments)

    if fail_count:
        print(
            f"WARNING: {fail_count} homebuilder constraint violation(s) detected — "
            "check Constraint Audit sheet.",
            file=sys.stderr,
        )

    wb.save(output_path)
    return output_path
